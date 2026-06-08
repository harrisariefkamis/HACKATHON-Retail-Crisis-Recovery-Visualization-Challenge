import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
from mlxtend.frequent_patterns import apriori, association_rules
import os

# ==========================================
# 1. LOAD & PREPARE DATA
# ==========================================
file_path = 'data_penjualan.xlsx'
try:
    df = pd.read_excel(file_path, sheet_name='Transaksi')
except Exception as e:
    print(f"Error: {e}")
    exit()

df['Tanggal'] = pd.to_datetime(df['tgl_transaksi'])

# Agregasi harian awal
daily_df = df.groupby(['tgl_transaksi', 'kode_produk', 'nama_produk'])['total_nilai'].sum().reset_index()
daily_df = daily_df.sort_values(by=['kode_produk', 'tgl_transaksi'])

# ==========================================
# 2. PERHITUNGAN MA & IDENTIFIKASI TREN
# ==========================================
window = 3
daily_df['MA'] = daily_df.groupby('kode_produk')['total_nilai'].transform(lambda x: x.rolling(window=window).mean())

# Tentukan tren naik
daily_df['Is_Rising'] = daily_df.groupby('kode_produk')['MA'].diff() > 0
daily_df['Trend_Session'] = (daily_df['Is_Rising'] != daily_df.groupby('kode_produk')['Is_Rising'].shift()).groupby(daily_df['kode_produk']).cumsum()

# Hitung durasi kenaikan berurutan
def count_consecutive(group):
    group['Consecutive_Rise'] = group.groupby('Trend_Session').cumcount() + 1
    group.loc[group['Is_Rising'] == False, 'Consecutive_Rise'] = 0
    return group

daily_df = daily_df.groupby('kode_produk', group_keys=False).apply(count_consecutive)

# ==========================================
# 3. NORMALISASI (BASE 100)
# ==========================================
def normalize_to_base_100(group):
    # Mengambil nilai MA pertama yang tersedia untuk normalisasi
    first_val = group['MA'].iloc[0]
    if pd.isna(first_val) or first_val == 0:
        # Jika nilai pertama NaN (karena window MA) atau 0, cari nilai valid pertama
        valid_vals = group['MA'].dropna()
        first_val = valid_vals.iloc[0] if not valid_vals.empty else 1
    
    group['Normalized'] = (group['MA'] / first_val) * 100
    return group

daily_df = daily_df.groupby('kode_produk', group_keys=False).apply(normalize_to_base_100)

# ==========================================
# 4. PERHITUNGAN GROWTH % (OPSI A)
# ==========================================
rising_sessions = daily_df[daily_df['Is_Rising'] == True].copy()

growth_per_session = rising_sessions.groupby(['kode_produk', 'nama_produk', 'Trend_Session']).agg(
    Growth_Pct=('MA', lambda x: (x.iloc[-1] / x.iloc[0] - 1) * 100),
    Max_Consecutive=('Consecutive_Rise', 'max')
).reset_index()

# Filter minimal tren 12 hari
target_rentetan = 12
mask_12_days = growth_per_session['Max_Consecutive'] >= target_rentetan
filtered_growth = growth_per_session[mask_12_days].copy()

# Ambil growth tertinggi per produk
final_growth = filtered_growth.groupby(['kode_produk', 'nama_produk'])['Growth_Pct'].max().reset_index()

# Merge dengan total penjualan riil
total_sales = df.groupby('kode_produk')['total_nilai'].sum().reset_index()
final_report = pd.merge(final_growth, total_sales, on='kode_produk')
final_report = final_report.sort_values(by='Growth_Pct', ascending=False)

# ==========================================
# 5. DISPLAY HASIL
# ==========================================
print(f"\nREKAPITULASI PRODUK (MINIMAL TREN {target_rentetan} HARI)")
print("-" * 90)

display_final = final_report.copy()
display_final.columns = ['Kode Produk', 'Nama Produk', 'Growth %', 'Total Penjualan']
display_final['Growth %'] = display_final['Growth %'].round(2) #.apply(lambda x: f"{x:.2f}%")
display_final['Total Penjualan'] = display_final['Total Penjualan'].astype(int) #.apply(lambda x: f"{x:,.0f}")

print(display_final.to_string(index=False))

# ==========================================
# 5B. APRIORI / PRODUCT PACKAGING
# ==========================================

print("\nMEMPROSES APRIORI ANALYSIS...")

# =====================================================
# BENTUK TRANSACTION MATRIX
# =====================================================

basket = (
    df.groupby(['nomor_struk', 'nama_produk'])['jumlah_terjual']
    .sum()
    .unstack(fill_value=0)
)

# Binary encode lebih cepat daripada map(lambda)
basket = (basket > 0).astype(int)

# =====================================================
# APRIORI
# =====================================================

frequent_itemsets = apriori(
    basket,
    min_support=0.01,
    use_colnames=True
)

# =====================================================
# ASSOCIATION RULES
# =====================================================

rules = association_rules(
    frequent_itemsets,
    metric='lift',
    min_threshold=1
)

# =====================================================
# FILTER: HARUS MENGANDUNG RISING STAR
# =====================================================

# Ambil daftar nama produk rising star
rising_star_products = set(final_report['nama_produk'])

# Fungsi cek apakah salah satu item ada di rising star
def contains_rising_star(itemset):
    return any(item in rising_star_products for item in itemset)

# Filter antecedents atau consequents
rules = rules[
    rules['antecedents'].apply(contains_rising_star) |
    rules['consequents'].apply(contains_rising_star)
].copy()


# Lift minimal > 2
rules = rules[rules['lift'] >= 2]

# Optional tambahan (recommended)
# rules = rules[rules['confidence'] >= 0.3]

# =====================================================
# SORTING
# =====================================================
rules = rules.sort_values(
    by=['lift', 'support', 'confidence'],
    ascending=[False, False, False]
)

# =====================================================
# HITUNG JUMLAH TRANSAKSI
# =====================================================
total_transactions = df['nomor_struk'].nunique()

rules['Jumlah_Transaksi_Rule'] = (
    rules['support'] * total_transactions
).round(0).astype(int)


# =====================================================
# FORMAT HASIL
# =====================================================

packaging_result = pd.DataFrame({
    'Jika Membeli': rules['antecedents'].apply(lambda x: ', '.join(sorted(x, reverse=True))),
    'Maka Membeli': rules['consequents'].apply(lambda x: ', '.join(sorted(x, reverse=True))),
    'Jumlah Invoice': rules['Jumlah_Transaksi_Rule'],
    'Support': rules['support'].round(2),
    'Confidence': rules['confidence'].round(2),
    'Lift': rules['lift'].round(2)
})

packaging_result = packaging_result.sort_values(
    by=['Lift', 'Support', 'Confidence'],
    ascending=[False, False, False]
)

print("\nTop 10 Packaging Recommendation:")
print(packaging_result.head(100))

# ===================
# 6. EXPORT KE EXCEL 
# ===================

output_file = 'retail-insight.xlsx'
sheet_name = 'Rising Star'

# Jika file sudah ada
if os.path.exists(output_file):

    with pd.ExcelWriter(
        output_file,
        engine='openpyxl',
        mode='a',
        if_sheet_exists='replace'
    ) as writer:

        # ======================================
        # SHEET RISING STAR
        # ======================================

        display_final.to_excel(
            writer,
            sheet_name='Rising Star',
            index=False
        )

        # ======================================
        # SHEET PACKAGING
        # ======================================

        packaging_result.to_excel(
            writer,
            sheet_name='Potential Packaging',
            index=False
        )
# Jika file belum ada
else:
    with pd.ExcelWriter(
        output_file,
        engine='openpyxl'
    ) as writer:

        display_final.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )

        packaging_result.to_excel(
            writer,
            sheet_name='Potential Packaging',
            index=False
        )
# ==========================================
# 7. FORMAT EXCEL
# ==========================================

workbook = load_workbook(output_file)
worksheet = workbook[sheet_name]

# Bold header
for cell in worksheet[1]:
    cell.font = Font(bold=True)

# Freeze pane
worksheet.freeze_panes = 'A2'

# Auto width + format angka
for column_cells in worksheet.columns:

    max_length = 0
    column_letter = get_column_letter(column_cells[0].column)

    for cell in column_cells:

        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass

    worksheet.column_dimensions[column_letter].width = max_length + 3

# Format numerik
for row in worksheet.iter_rows(min_row=2):

    # Growth %
    row[2].number_format = '0.00'

    # Total Penjualan
    row[3].number_format = '#,##0'

workbook.save(output_file)

print(f"\nSheet '{sheet_name}' berhasil dibuat.")


# ==================================================
# 8. VISUALISASI DENGAN WARNA CUSTOM PER PERINGKAT
# ==================================================

# Filter data hanya untuk produk yang masuk dalam report final
plot_df = daily_df[daily_df['kode_produk'].isin(final_report['kode_produk'])].copy()

if not plot_df.empty:

    # ============================================================
    # A. SPESIFIKASI FIGURE
    # ============================================================
    fig = plt.figure(figsize=(15, 8), dpi=100)
    ax = fig.add_subplot(111)

    # ============================================================
    # B. PENGATURAN WARNA CUSTOM BERDASARKAN PERINGKAT
    # ============================================================

    # Urutkan berdasarkan growth tertinggi
    sorted_report = final_report.sort_values(
        by='Growth_Pct',
        ascending=False
    )

    # Palet warna custom
    custom_palette = [
        '#FFD700',  # Gold
        '#C0C0C0',  # Silver
        '#CD7F32',  # Bronze
        '#2ecc71',  # Emerald Green
        '#3498db',  # Blue
        '#9b59b6',  # Purple
        '#e74c3c',  # Red
        '#34495e',  # Dark Blue Grey
    ]

    default_color = '#95a5a6'

    # Mapping warna & ranking
    color_mapping = {}
    rank_mapping = {}

    for i, row in enumerate(sorted_report.itertuples()):

        kode_produk = row.kode_produk

        color_mapping[kode_produk] = (
            custom_palette[i]
            if i < len(custom_palette)
            else default_color
        )

        rank_mapping[kode_produk] = i + 1

    # ============================================================
    # C. TOP 3 PRODUK BERDASARKAN TOTAL PENJUALAN
    # ============================================================

    top3_sales = (
        df.groupby(['kode_produk', 'nama_produk'])['total_nilai']
          .sum()
          .reset_index()
          .sort_values(by='total_nilai', ascending=False)
          .head(3)
    )

    top3_codes = top3_sales['kode_produk'].tolist()

    top3_plot_df = daily_df[
        daily_df['kode_produk'].isin(top3_codes)
    ].copy()

    # ============================================================
    # D. PLOT TOP 3 SALES (ABU-ABU)
    # ============================================================

    grey_colors = [
        '#B0B0B0',
        '#909090',
        '#707070'
    ]

    for idx, (kode_produk, group) in enumerate(
        top3_plot_df.groupby('kode_produk')
    ):

        nama_produk = group['nama_produk'].iloc[0]

        grey_color = (
            grey_colors[idx]
            if idx < len(grey_colors)
            else '#808080'
        )

        ax.plot(
            group['tgl_transaksi'],
            group['Normalized'],
            linestyle='--',
            linewidth=2,
            marker='o',
            markersize=3,
            color=grey_color,
            alpha=0.7,
            label=f"Top Sales: {nama_produk}"
        )

    # ============================================================
    # E. PLOT RISING STAR
    # ============================================================

    for kode_produk, group in plot_df.groupby('kode_produk'):

#        rank = rank_mapping.get(kode_produk, 999)
#
#        # Hanya tampilkan Rank 1
#        if rank != 1:
#            continue
#
#        nama_produk = group['nama_produk'].iloc[0]
#
#        line_color = color_mapping.get(
#            kode_produk,
#            default_color   
#        )
    
        nama_produk = group['nama_produk'].iloc[0]

        line_color = color_mapping.get(
            kode_produk,
            default_color
        )

        rank = rank_mapping.get(
            kode_produk,
            '?'
        )


        label_with_rank = f"Rank {rank}: {nama_produk}"

        ax.plot(
            group['tgl_transaksi'],
            group['Normalized'],
            marker='o',
            markersize=4,
            linewidth=2.5,
            color=line_color,
            label=label_with_rank
        )

    # ============================================================
    # F. TITLE & LABEL
    # ============================================================

    font_title = {
        'family': 'sans-serif',
        'color': 'black',
        'weight': 'bold',
        'size': 16
    }

    font_label = {
        'family': 'sans-serif',
        'weight': 'normal',
        'size': 12
    }

    ax.set_title(
        'ANALISIS PERTUMBUHAN RELATIF PRODUK RISING STAR\n'
        '(Dengan Benchmark Top 3 Total Penjualan)',
        fontdict=font_title,
        pad=20
    )

    ax.set_xlabel(
        'Periode Tanggal',
        fontdict=font_label,
        labelpad=10
    )

    ax.set_ylabel(
        'Indeks Pertumbuhan (Base 100)',
        fontdict=font_label,
        labelpad=10
    )

    # ============================================================
    # G. GRID & BASELINE
    # ============================================================

    ax.grid(
        True,
        linestyle='--',
        linewidth=0.5,
        alpha=0.5
    )

    ax.axhline(
        y=100,
        color='black',
        linestyle='-',
        linewidth=1,
        alpha=0.5
    )

    # ============================================================
    # H. FORMAT AXIS
    # ============================================================

    plt.xticks(
        rotation=45,
        ha='right',
        fontsize=10
    )

    plt.yticks(fontsize=10)

    # ============================================================
    # I. SORT LEGEND BERDASARKAN RANK
    # ============================================================

    handles, labels = ax.get_legend_handles_labels()

    # Pisahkan Top Sales & Rising Star
    top_sales_items = []
    rising_items = []

    for h, l in zip(handles, labels):

        if l.startswith('Top Sales'):
            top_sales_items.append((h, l))
        else:
            rising_items.append((h, l))

    # Sort rising star berdasarkan ranking
    rising_items = sorted(
        rising_items,
        key=lambda x: int(
            x[1].split(':')[0].split()[1]
        )
    )

    # Gabungkan kembali
    final_legend = top_sales_items + rising_items

    final_handles = [x[0] for x in final_legend]
    final_labels = [x[1] for x in final_legend]

    # ============================================================
    # J. LEGEND
    # ============================================================

    ax.legend(
        final_handles,
        final_labels,
        title="Kategori Produk",
        title_fontsize=12,
        fontsize=10,
        bbox_to_anchor=(1.02, 1),
        loc='upper left',
        borderaxespad=0,
        frameon=True,
        shadow=True
    )

    # ============================================================
    # K. LAYOUT & SAVE
    # ============================================================

    plt.tight_layout()

    plt.savefig(
        'rising_star_index.png',
        bbox_inches='tight'
    )

    print(
        "\nGrafik detail disimpan sebagai "
        "'rising_star_index.png'"
    )

    # plt.show()

else:

    print(
        "\nTidak ada data untuk di-plot "
        "(mungkin tidak ada produk memenuhi kriteria)."
    )
    

# ============================================================
# 9. VISUALISASI NILAI PENJUALAN ASLI
# ============================================================

fig2 = plt.figure(figsize=(15, 8), dpi=100)
ax2 = fig2.add_subplot(111)

# ============================================================
# A. PLOT TOP 3 SALES
# ============================================================

for idx, (kode_produk, group) in enumerate(
    top3_plot_df.groupby('kode_produk')
):

    nama_produk = group['nama_produk'].iloc[0]

    grey_color = (
        grey_colors[idx]
        if idx < len(grey_colors)
        else '#808080'
    )

    ax2.plot(
        group['tgl_transaksi'],
        group['total_nilai'],
        linestyle='--',
        linewidth=2,
        marker='o',
        markersize=3,
        color=grey_color,
        alpha=0.7,
        label=f"Top Sales: {nama_produk}"
    )

# ============================================================
# B. PLOT RISING STAR BERDASARKAN NILAI ASLI
# ============================================================

for kode_produk, group in plot_df.groupby('kode_produk'):

    nama_produk = group['nama_produk'].iloc[0]

    line_color = color_mapping.get(
        kode_produk,
        default_color
    )

    rank = rank_mapping.get(
        kode_produk,
        '?'
    )
#    rank = rank_mapping.get(kode_produk, 999)
#
#    # Hanya tampilkan Rank 1
#    if rank != 1:
#        continue
#
#    nama_produk = group['nama_produk'].iloc[0]
#
#    line_color = color_mapping.get(
#        kode_produk,
#        default_color
#    )
    
    label_with_rank = f"Rank {rank}: {nama_produk}"

    ax2.plot(
        group['tgl_transaksi'],
        group['total_nilai'],
        marker='o',
        markersize=4,
        linewidth=2.5,
        color=line_color,
        label=label_with_rank
    )

# ============================================================
# C. TITLE & LABEL
# ============================================================

ax2.set_title(
    'ANALISIS NILAI PENJUALAN PRODUK RISING STAR\n'
    '(Nilai Penjualan Asli)',
    fontdict=font_title,
    pad=20
)

ax2.set_xlabel(
    'Periode Tanggal',
    fontdict=font_label,
    labelpad=10
)

ax2.set_ylabel(
    'Total Nilai Penjualan',
    fontdict=font_label,
    labelpad=10
)

# ============================================================
# D. GRID
# ============================================================

ax2.grid(
    True,
    linestyle='--',
    linewidth=0.5,
    alpha=0.5
)

# ============================================================
# E. FORMAT AXIS
# ============================================================

plt.xticks(
    rotation=45,
    ha='right',
    fontsize=10
)

plt.yticks(fontsize=10)

# ============================================================
# F. SORT LEGEND
# ============================================================

handles2, labels2 = ax2.get_legend_handles_labels()

top_sales_items2 = []
rising_items2 = []

for h, l in zip(handles2, labels2):

    if l.startswith('Top Sales'):
        top_sales_items2.append((h, l))
    else:
        rising_items2.append((h, l))

rising_items2 = sorted(
    rising_items2,
    key=lambda x: int(
        x[1].split(':')[0].split()[1]
    )
)

final_legend2 = top_sales_items2 + rising_items2

final_handles2 = [x[0] for x in final_legend2]
final_labels2 = [x[1] for x in final_legend2]

# ============================================================
# G. LEGEND
# ============================================================

ax2.legend(
    final_handles2,
    final_labels2,
    title="Kategori Produk",
    title_fontsize=12,
    fontsize=10,
    bbox_to_anchor=(1.02, 1),
    loc='upper left',
    borderaxespad=0,
    frameon=True,
    shadow=True
)

# ============================================================
# H. LAYOUT & SAVE
# ============================================================

plt.tight_layout()

plt.savefig(
    'rising_star_actual.png',
    bbox_inches='tight'
)

print(
    "\nGrafik nilai penjualan disimpan sebagai "
    "'rising_star_actual.png'"
)

# plt.show()