"""
╔══════════════════════════════════════════════════════════════════════╗
║  DQLab Hackathon — Retail Crisis & Recovery Visualization Challenge  ║
║  Script : solusi-retail.py                                           ║
║  Output : retail_insight.xlsx | rising_star_index.png               ║
║           rising_star_actual.png                                     ║
╚══════════════════════════════════════════════════════════════════════╝

Menjalankan: python solusi-retail.py
"""

# ─── IMPORTS ────────────────────────────────────────────────────────────────
import warnings
warnings.filterwarnings('ignore')

import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.lines import Line2D
from mlxtend.frequent_patterns import apriori, association_rules
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ─── CONFIGURATION ──────────────────────────────────────────────────────────
INPUT_FILE   = 'data_penjualan.csv'          # ganti jika nama file berbeda
OUTPUT_EXCEL = 'retail_insight.xlsx'
OUTPUT_INDEX = 'rising_star_index.png'
OUTPUT_ACTUAL= 'rising_star_actual.png'
MIN_CONSEC   = 12                            # minimum consecutive rising days
MA_WINDOW    = 3                             # moving average window
MIN_SUPPORT  = 0.01
LIFT_METRIC  = 'lift'
MIN_LIFT_THRESHOLD = 1.0
MIN_LIFT_FILTER    = 2.0

print("=" * 65)
print("  DQLab Hackathon — Retail Crisis & Recovery")
print("=" * 65)

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  1. LOAD DATA                                                        ║
# ╚══════════════════════════════════════════════════════════════════════╝
print("\n[1/5] Loading data...")

# Coba baca file dari working directory atau path absolut
if not os.path.exists(INPUT_FILE):
    INPUT_FILE = '/mnt/user-data/uploads/data_penjualan.csv'

df = pd.read_csv(INPUT_FILE, sep=';')
df['tgl_transaksi'] = pd.to_datetime(df['tgl_transaksi'], format='%d-%m-%Y')
df['kode_produk']   = df['kode_produk'].str.strip().str.replace(' ', '', regex=False)
df['nama_produk']   = df['nama_produk'].str.strip()

print(f"  Rows       : {len(df):,}")
print(f"  Date range : {df['tgl_transaksi'].min().date()} → {df['tgl_transaksi'].max().date()}")
print(f"  SKUs       : {df['kode_produk'].nunique()}")
print(f"  Invoices   : {df['nomor_struk'].nunique():,}")

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  2. RISING STAR ANALYSIS                                             ║
# ╚══════════════════════════════════════════════════════════════════════╝
print("\n[2/5] Computing Rising Star...")

# 2a. Aggregate daily total_nilai per product
daily_prod = (df.groupby(['kode_produk', 'nama_produk', 'tgl_transaksi'])['total_nilai']
              .sum().reset_index())

all_dates = pd.date_range(df['tgl_transaksi'].min(), df['tgl_transaksi'].max(), freq='D')

rising_results = []

# Precompute total penjualan per produk sekali untuk menghindari df[df['kode_produk']==...] berulang
_total_by_kode = (
    df.groupby('kode_produk', sort=False)['total_nilai']
      .sum()
      .astype(float)
)

for (kode, nama), grp in daily_prod.groupby(['kode_produk', 'nama_produk']):

    # Fill missing dates dengan reindex sekali per SKU
    ts = (grp.set_index('tgl_transaksi')[['total_nilai']]
          .reindex(all_dates, fill_value=0)
          .reset_index()
          .rename(columns={'index': 'tgl_transaksi'}))

    # Moving Average (3 hari)
    vals = ts['total_nilai'].to_numpy(dtype=float)
    cumsum = np.cumsum(vals)

    # MA_window=3, hitung cepat via rolling min_periods=1
    # ma[t] = mean(vals[max(0,t-2):t+1])
    ma = np.empty_like(vals, dtype=float)
    for t in range(len(vals)):
        start = 0 if t < MA_WINDOW - 1 else t - (MA_WINDOW - 1)
        window_sum = cumsum[t] - (cumsum[start - 1] if start > 0 else 0.0)
        ma[t] = window_sum / (t - start + 1)

    # rising flag: ma[t] > ma[t-1]
    rising = np.zeros(len(vals), dtype=bool)
    rising[1:] = ma[1:] > ma[:-1]

    # Longest consecutive TRUE (vector-friendly scan)
    max_consec = 0
    best_start_i = 0
    best_end_i = 0
    cur_consec = 0
    cur_start_i = 0

    for i in range(len(rising)):
        if rising[i]:
            if cur_consec == 0:
                cur_start_i = i
            cur_consec += 1
            if cur_consec > max_consec:
                max_consec = cur_consec
                best_start_i = cur_start_i
                best_end_i = i
        else:
            cur_consec = 0

    if max_consec >= MIN_CONSEC:
        ma_start = ma[best_start_i]
        ma_end   = ma[best_end_i]

        growth_pct = ((ma_end / ma_start) - 1) * 100 if ma_start > 0 else 0
        total_penjualan = float(_total_by_kode.get(kode, 0.0))

        rising_results.append({
            'Kode Produk'     : kode,
            'Nama Produk'     : nama,
            'Growth %'        : round(growth_pct, 2),
            'Total Penjualan' : int(total_penjualan),
            '_max_consec'     : max_consec,
            # simpan ts hanya untuk produk yang lolos threshold (tetap sesuai kebutuhan chart)
        '_ts'             : ts.assign(ma=ma),
        })



# Sort by Growth % descending
rising_results.sort(key=lambda x: x['Growth %'], reverse=True)

# Assign rank
for i, r in enumerate(rising_results):
    r['_rank'] = i + 1

rising_df = pd.DataFrame(rising_results)[['Kode Produk', 'Nama Produk',
                                           'Growth %', 'Total Penjualan']]
rising_kodes = rising_df['Kode Produk'].tolist()

print(f"  Rising Star products found: {len(rising_df)}")
for _, row in rising_df.iterrows():
    print(f"    → {row['Kode Produk']:10s}  {row['Nama Produk']:<35s}  "
          f"Growth: {row['Growth %']:,.2f}%  Total: Rp {row['Total Penjualan']:,}")

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  3. POTENTIAL PACKAGING (Apriori)                                    ║
# ╚══════════════════════════════════════════════════════════════════════╝
print("\n[3/5] Computing Potential Packaging (Apriori)...")

# Build basket matrix: nomor_struk × kode_produk (True/False)
basket = (df.groupby(['nomor_struk', 'kode_produk'])['jumlah_terjual']
          .sum().unstack(fill_value=0) > 0)

total_invoices = len(basket)
print(f"  Total invoices for basket: {total_invoices:,}")

# Frequent itemsets
freq_items = apriori(basket, min_support=MIN_SUPPORT, use_colnames=True)
print(f"  Frequent itemsets found: {len(freq_items)}")

# Association rules
rules = association_rules(freq_items, metric=LIFT_METRIC,
                          min_threshold=MIN_LIFT_THRESHOLD)
print(f"  Total rules (lift≥{MIN_LIFT_THRESHOLD}): {len(rules)}")

# Kode ke nama mapping
kode_nama = df[['kode_produk','nama_produk']].drop_duplicates().set_index('kode_produk')['nama_produk'].to_dict()

def items_to_str(frozen):
    """Convert frozenset of kode_produk to human-readable comma list."""
    names = [kode_nama.get(k, k) for k in sorted(frozen)]
    return ', '.join(names)

def has_rising_star(frozen):
    return any(k in rising_kodes for k in frozen)

# Filter rules
filtered = rules[
    (rules['antecedents'].apply(has_rising_star) |
     rules['consequents'].apply(has_rising_star)) &
    (rules['lift'] >= MIN_LIFT_FILTER)
].copy()

# Calculate Jumlah Invoice (support count)
filtered['Jumlah Invoice'] = (filtered['support'] * total_invoices).round(0).astype(int)
filtered['Jika Membeli']   = filtered['antecedents'].apply(items_to_str)
filtered['Maka Membeli']   = filtered['consequents'].apply(items_to_str)
filtered['Support']        = filtered['support'].round(4)
filtered['Confidence']     = filtered['confidence'].round(4)
filtered['Lift']           = filtered['lift'].round(4)

# Sort: Lift DESC, Support DESC, Confidence DESC
packaging_df = (filtered[['Jika Membeli', 'Maka Membeli', 'Jumlah Invoice',
                            'Support', 'Confidence', 'Lift']]
                .sort_values(['Lift','Support','Confidence'], ascending=False)
                .reset_index(drop=True))

print(f"  Filtered rules (has Rising Star & lift≥{MIN_LIFT_FILTER}): {len(packaging_df)}")
print(packaging_df.head(5).to_string(index=False))

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  4. EXPORT TO EXCEL                                                  ║
# ╚══════════════════════════════════════════════════════════════════════╝
print("\n[4/5] Exporting retail_insight.xlsx...")

with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
    rising_df.to_excel(writer, sheet_name='Rising Star',      index=False)
    packaging_df.to_excel(writer, sheet_name='Potential Packaging', index=False)

# Style the workbook
wb = openpyxl.load_workbook(OUTPUT_EXCEL)

# ── Styling helpers ──────────────────────────────────────────────────
HEADER_FILL_RS  = PatternFill('solid', fgColor='1F4E79')
HEADER_FILL_PP  = PatternFill('solid', fgColor='833C00')
HEADER_FONT     = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
EVEN_FILL       = PatternFill('solid', fgColor='EBF3FB')
ODD_FILL        = PatternFill('solid', fgColor='FFFFFF')
THIN_BORDER     = Border(
    left  =Side(style='thin'),  right =Side(style='thin'),
    top   =Side(style='thin'),  bottom=Side(style='thin'),
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_ALIGN   = Alignment(horizontal='left',   vertical='center')

def style_sheet(ws, header_fill, num_fmt_map=None):
    """Apply header styling and alternating row colors."""
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.font      = HEADER_FONT
        cell.fill      = header_fill
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for cell in row:
            cell.fill      = fill
            cell.border    = THIN_BORDER
            cell.alignment = CENTER_ALIGN

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

    ws.row_dimensions[1].height = 25

# Style Rising Star sheet
ws_rs = wb['Rising Star']
style_sheet(ws_rs, HEADER_FILL_RS)

# Format numbers in Rising Star
for row in ws_rs.iter_rows(min_row=2):
    kode_cell      = row[0];  kode_cell.alignment = CENTER_ALIGN
    nama_cell      = row[1];  nama_cell.alignment = LEFT_ALIGN
    growth_cell    = row[2]
    total_cell     = row[3]
    growth_cell.number_format = '#,##0.00'
    total_cell.number_format  = '#,##0'

# Style Potential Packaging sheet
ws_pp = wb['Potential Packaging']
style_sheet(ws_pp, HEADER_FILL_PP)

for row in ws_pp.iter_rows(min_row=2):
    row[0].alignment = LEFT_ALIGN   # Jika Membeli
    row[1].alignment = LEFT_ALIGN   # Maka Membeli
    row[2].number_format = '#,##0'  # Jumlah Invoice
    row[3].number_format = '#,##0.0000'  # Support
    row[4].number_format = '#,##0.0000'  # Confidence
    row[5].number_format = '#,##0.0000'  # Lift

wb.save(OUTPUT_EXCEL)
print(f"  ✅  Saved → {OUTPUT_EXCEL}")

# ╔══════════════════════════════════════════════════════════════════════╗
# ║  5. VISUALIZATIONS                                                   ║
# ╚══════════════════════════════════════════════════════════════════════╝
print("\n[5/5] Generating visualizations...")

# ── Prepare daily MA per product ─────────────────────────────────────
def get_daily_ma(kode_list, daily_prod=daily_prod, all_dates=all_dates, ma_window=MA_WINDOW):
    """Return dict of {kode: pd.Series(MA, index=date)} for given kode list.

    Optimized: avoid calling groupby repeatedly and avoid per-SKU reindex if possible.
    """
    kode_set = set(kode_list)

    # Only keep needed SKUs to reduce work
    dp = daily_prod[daily_prod['kode_produk'].isin(kode_set)]

    out = {}
    # Group once over filtered dataset
    for (kode, _nama), grp in dp.groupby(['kode_produk', 'nama_produk']):
        ts = (
            grp.set_index('tgl_transaksi')[['total_nilai']]
            .reindex(all_dates, fill_value=0)
            .reset_index()
            .rename(columns={'index': 'tgl_transaksi'})
        )
        ts['ma'] = ts['total_nilai'].rolling(window=ma_window, min_periods=1).mean()
        out[kode] = ts.set_index('tgl_transaksi')['ma']

    return out

# Top 3 products by total penjualan
top3 = (df.groupby(['kode_produk','nama_produk'])['total_nilai']
        .sum().reset_index()
        .sort_values('total_nilai', ascending=False)
        .head(3))
top3_kodes  = top3['kode_produk'].tolist()
top3_nama   = dict(zip(top3['kode_produk'], top3['nama_produk']))
rs_kode_nama= dict(zip(rising_df['Kode Produk'], rising_df['Nama Produk']))

all_kodes_needed = list(set(rising_kodes + top3_kodes))
ma_dict = get_daily_ma(all_kodes_needed)

# Colour palettes (matching PDF style)
TOP3_COLORS = ['#595959', '#7F7F7F', '#A6A6A6']   # gray shades for top 3 (dashed)
RS_COLORS   = ['#FFC000', '#ED7D31', '#5B9BD5',
               '#70AD47', '#FF4747', '#BC8CFF']      # distinct for rising stars

# ════════════════════════════════════════════════════════════════════
# CHART A — rising_star_index.png  (Normalisasi Base 100)
# ════════════════════════════════════════════════════════════════════
fig_idx, ax_idx = plt.subplots(figsize=(14, 6))
fig_idx.patch.set_facecolor('white')
ax_idx.set_facecolor('white')

legend_handles = []

# Plot Top 3 (dashed)
for i, kode in enumerate(top3_kodes):
    if kode not in ma_dict:
        continue
    s   = ma_dict[kode]
    s0  = s.iloc[0] if s.iloc[0] != 0 else 1
    idx_s = (s / s0) * 100

    color = TOP3_COLORS[i % len(TOP3_COLORS)]
    ax_idx.plot(idx_s.index, idx_s.values, linestyle='--',
                color=color, linewidth=1.2, alpha=0.85)
    legend_handles.append(
        Line2D([0],[0], linestyle='--', color=color, linewidth=1.2,
               label=f"Top Sales: {top3_nama[kode]}")
    )

# Plot Rising Stars (solid)
for i, r in enumerate(rising_results):
    kode = r['Kode Produk']
    if kode not in ma_dict:
        continue
    s   = ma_dict[kode]
    s0  = s.iloc[0] if s.iloc[0] != 0 else 1
    idx_s = (s / s0) * 100

    color = RS_COLORS[i % len(RS_COLORS)]
    ax_idx.plot(idx_s.index, idx_s.values, linestyle='-',
                color=color, linewidth=2.0)
    legend_handles.append(
        Line2D([0],[0], linestyle='-', color=color, linewidth=2.0,
               label=f"Rank {r['_rank']}: {r['Nama Produk']}")
    )

# Formatting
ax_idx.set_title(
    "ANALISIS PERTUMBUHAN RELATIF PRODUK RISING STAR\n"
    "(Dengan Benchmark Top 3 Total Penjualan)",
    fontsize=13, fontweight='bold', color='black', pad=12
)
ax_idx.set_ylabel("Indeks Pertumbuhan (Base 100)", fontsize=10, color='black')
ax_idx.set_xlabel("Periode Tanggal", fontsize=10, color='black')
ax_idx.tick_params(axis='x', rotation=45, labelsize=8)
ax_idx.tick_params(axis='y', labelsize=9)
ax_idx.grid(True, linestyle='--', alpha=0.5, color='#CCCCCC')
ax_idx.spines[['top','right']].set_visible(False)

legend = ax_idx.legend(
    handles=legend_handles,
    title='Kategori Produk',
    title_fontsize=9,
    fontsize=8,
    loc='upper left',
    bbox_to_anchor=(1.01, 1),
    borderaxespad=0,
    frameon=True,
    framealpha=0.9,
)
legend.get_title().set_fontweight('bold')

plt.tight_layout()
fig_idx.savefig(OUTPUT_INDEX, dpi=150, bbox_inches='tight', facecolor='white')
plt.close(fig_idx)
print(f"  ✅  Saved → {OUTPUT_INDEX}")

# ════════════════════════════════════════════════════════════════════
# CHART B — rising_star_actual.png  (Nilai Aktual MA)
# Shows: Top 3 products + Rank 1 Rising Star
# ════════════════════════════════════════════════════════════════════
fig_act, ax_act = plt.subplots(figsize=(14, 6))
fig_act.patch.set_facecolor('white')
ax_act.set_facecolor('white')

legend_handles_act = []

# Plot Top 3 (dashed) — actual MA values
for i, kode in enumerate(top3_kodes):
    if kode not in ma_dict:
        continue
    s     = ma_dict[kode]
    color = TOP3_COLORS[i % len(TOP3_COLORS)]
    ax_act.plot(s.index, s.values, linestyle='--',
                color=color, linewidth=1.2, alpha=0.85)
    legend_handles_act.append(
        Line2D([0],[0], linestyle='--', color=color, linewidth=1.2,
               label=f"Top Sales: {top3_nama[kode]}")
    )

# Plot Rank 1 Rising Star only — solid
if rising_results:
    top_rs = rising_results[0]
    kode   = top_rs['Kode Produk']
    if kode in ma_dict:
        s     = ma_dict[kode]
        color = RS_COLORS[0]
        ax_act.plot(s.index, s.values, linestyle='-',
                    color=color, linewidth=2.0)
        legend_handles_act.append(
            Line2D([0],[0], linestyle='-', color=color, linewidth=2.0,
                   label=f"Rank 1: {top_rs['Nama Produk']}")
        )

# Formatting
ax_act.set_title(
    "ANALISIS NILAI PENJUALAN PRODUK RISING STAR\n"
    "(Nilai Penjualan Asli)",
    fontsize=13, fontweight='bold', color='black', pad=12
)
ax_act.set_ylabel("Total Nilai Penjualan", fontsize=10, color='black')
ax_act.set_xlabel("Periode Tanggal", fontsize=10, color='black')
ax_act.tick_params(axis='x', rotation=45, labelsize=8)
ax_act.tick_params(axis='y', labelsize=9)
ax_act.yaxis.set_major_formatter(mticker.ScalarFormatter(useMathText=True))
ax_act.ticklabel_format(axis='y', style='sci', scilimits=(7, 7))
ax_act.grid(True, linestyle='--', alpha=0.5, color='#CCCCCC')
ax_act.spines[['top','right']].set_visible(False)

legend_act = ax_act.legend(
    handles=legend_handles_act,
    title='Kategori Produk',
    title_fontsize=9,
    fontsize=8,
    loc='upper left',
    bbox_to_anchor=(1.01, 1),
    borderaxespad=0,
    frameon=True,
    framealpha=0.9,
)
legend_act.get_title().set_fontweight('bold')

plt.tight_layout()
fig_act.savefig(OUTPUT_ACTUAL, dpi=150, bbox_inches='tight', facecolor='white')
plt.close(fig_act)
print(f"  ✅  Saved → {OUTPUT_ACTUAL}")

# ─── FINAL SUMMARY ──────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("  PIPELINE COMPLETE — Output Files:")
for f in [OUTPUT_EXCEL, OUTPUT_INDEX, OUTPUT_ACTUAL]:
    size = os.path.getsize(f) / 1024 if os.path.exists(f) else 0
    print(f"    {f:<35}  ({size:.0f} KB)")
print("=" * 65)
print("  DQLab Hackathon — solusi-retail.py  ✔")
print("=" * 65)
